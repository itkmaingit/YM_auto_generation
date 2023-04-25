import getpass
import os
import pickle

import init
import openpyxl as opl
import requests
from tqdm import tqdm


def emotion(name, client):
    filepath = f"{os.path.expanduser('~')}\\works\\{client}\\{name}"
    print("")
    print("---------------emotion.py-------------------")


    # EXCELファイルを開く．excel_sizeには最大列の番号を代入．
    wb = opl.load_workbook(
        f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\daihon_complete.xlsx"
    )
    ws = wb["Sheet"]
    excel_size = ws.max_row

    if os.path.isfile(f"{filepath}\\emotion.pickle"):
        with open(f"{filepath}\\emotion.pickle", "rb") as f:
            j = pickle.load(f)
            j = 1
    else:
        j = 1

    for i in tqdm(range(j, excel_size + 1)):

        # 送信パラメータ
        data = {"api_key": API_key, "text": ws["B" + str(i)].value}

        try:
            # txt.pyで作成したtxtファイルを読み込んで，感情分析をする．
            with open(
                f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\txt_file\\"
                + str(i)
                + ".txt"
            ) as text:
                r = requests.post(API_Endpoint, data=data, files={"text_data": text})

            # 結果をresultsに代入しておく．
            results = r.json()

        # たまにEXCELの最大列の取得のミスがあるので，例外処理．
        except FileNotFoundError:
            print("終わったみたいです．")
            break

        try:
            # emotion_detailを取得．joy,anger等の辞書型オブジェクト．
            e_value = results["emotion_detail"]
            # e_valueの中で最大値を取得．

            emotion = max(e_value, key=e_value.get)

            # 各々の感情に対応する数値を代入していく．
            if "！？" in ws["B" + str(i)].value or "!?" in ws["B" + str(i)].value:
                ws["C" + str(i)].value = 4
            elif "？" in ws["B" + str(i)].value or "?" in ws["B" + str(i)].value:
                ws["C" + str(i)].value = 3
            elif "・・" in ws["B" + str(i)].value or ".." in ws["B" + str(i)].value:
                ws["C" + str(i)].value = 8
            elif emotion == "fear":
                ws["C" + str(i)].value = 5
            elif emotion == "anger":
                ws["C" + str(i)].value = 7
            elif emotion == "joy":
                ws["C" + str(i)].value = 2
            elif emotion == "love":
                ws["C" + str(i)].value = 1
            elif emotion == "sadness":
                ws["C" + str(i)].value = 6

        except TypeError:
            break

        except KeyError:
            print("今日はもう上限です！")
            with open(f"{filepath}\\emotion.pickle", "wb") as f:
                pickle.dump(i, f)
                print(f"{i}番目で終了しました！")
            break

    # 保存．
    wb.save(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon_completed.xlsx"
    )


if __name__ == "__main__":
    name, client = init.main()
    emotion(name, client)
