import getpass
import glob
import os
import docx
import init
import openpyxl as opl
from tqdm import tqdm


def main(name, client):
    print("")
    print("---------------docx_xlsx.py-------------------")

    filename = sorted(
        glob.glob(
            f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\{name}.docx"
        )
    )
    doc = docx.Document(filename[0])

    wb = opl.Workbook()
    ws = wb.active
    n = 0
    k = 0

    for para in tqdm(doc.paragraphs):
        if para.text.strip() == "":
            if n >= 500:
                break
            else:
                pass
        elif len(para.text) == 1:
            ws[f"A{str(n+k+1)}"].value = ws[f"A{str(n+k)}"].value
            ws[f"B{str(n+k+1)}"].value = para.text.replace(" ", "").replace("　", "")
            n += 1
        elif para.text[1] != ":":
            ws[f"A{str(n+k+1)}"].value = ws[f"A{str(n+k)}"].value
            ws[f"B{str(n+k+1)}"].value = para.text.replace(" ", "").replace("　", "")
            n += 1
        elif para.text[0] == "0":
            ws[f"A{str(n+k+1)}"].value = "霊夢"
            ws[f"B{str(n+k+1)}"].value = para.text[2:].replace(" ", "").replace("　", "")
            n += 1
        elif para.text[0] == "1":
            ws[f"A{str(n+k+1)}"].value = "魔理沙"
            ws[f"B{str(n+k+1)}"].value = para.text[2:].replace(" ", "").replace("　", "")
            n += 1
        elif para.text[0] == "2":
            ws[f"A{str(n+k+1)}"].value = "2人"
            ws[f"B{str(n+k+1)}"].value = para.text[2:].replace(" ", "").replace("　", "")
            n += 1
        elif para.text[0] == "3":
            ws[f"A{str(n+k+1)}"].value = "タイトル"
            ws[f"B{str(n+k+1)}"].value = para.text[2:].replace(" ", "").replace("　", "")
            n += 1

        elif para.text[0] == "4":
            ws[f"A{str(n+k+1)}"].value = "フェード"
            ws[f"B{str(n+k+1)}"].value = "楔楔楔むら"
            n += 1

        else:
            ws[f"A{str(n+k+1)}"].value = "霊夢"
            ws[f"B{str(n+k+1)}"].value = para.text[2:].replace(" ", "").replace("　", "")
            n += 1

    wb.save(f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\daihon.xlsx")


if __name__ == "__main__":
    name, client = init.main()
    main(name, client)
