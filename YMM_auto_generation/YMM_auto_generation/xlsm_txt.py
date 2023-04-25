import getpass
import os
import init
import openpyxl as opl
from tqdm import tqdm


def extract(name, client):
    print("")
    print("---------------xlsm_txt.py-------------------")

    wb = opl.load_workbook(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon_complete.xlsx"
    )
    ws = wb["Sheet"]
    max_row = ws.max_row

    n = 1
    lst = []

    with open(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon.txt",
        "w",
        encoding="utf-8",
    ) as wf:
        with tqdm(total=max_row) as pbar:
            while ws["A" + str(n)].value is not None:

                text = ws["B" + str(n)].value.replace("「", "『")
                text = text.replace("」", "』")
                lst.append(ws["A" + str(n)].value + "「" + text + "」\n")
                n += 1
                pbar.update(1)

        wf.writelines(lst)


if __name__ == "__main__":
    name, client = init.main()
    extract(name, client)
