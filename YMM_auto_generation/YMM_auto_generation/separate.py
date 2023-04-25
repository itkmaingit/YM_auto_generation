import getpass
import os
import init
import openpyxl as opl
from janome.tokenizer import Tokenizer
from tqdm import tqdm


def separate(name, client):
    """読み込んだxlsxファイルに対し、指定文字数以下になるように文章を分割して次のセルへ渡す関数。
    MAX_C=40と指定すれば、130文字の時は4つの文章に、50文字の時は2つの文章に大体半分になるように分割する。

    Args:
        MAX_C: 指定文字数
        s_list: 文章の単語リスト
        s_part1: 各単語の品詞名
        s_part2: 各単語の品詞細分類(格助詞とか副助詞とか)
        s_length: 形式定数，改行文字挿入位置判定に使用
        k: 初めにk+1とするから-1始まりのほうが都合が良い．
        l: 分割した後の行数をカウントする。3行に分割したなら+2したい
        n: 開始の列番号
        j: スライスを取る変数text[:j[0]],text[j[0]:j[1]]のような使い方をする
    """
    print("")
    print("---------------separate.py-------------------")

    # 定数設定．
    MAX_C = 37
    n = 1
    word_amount = 0
    t = Tokenizer()  # 形態素解析のためのオブジェクト．

    # フォルダ名/daihon.xlsxを読み込む
    wb = opl.load_workbook(
        f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\daihon.xlsx"
    )
    ws = wb.active
    max_row = ws.max_row

    # 新しいexcelファイルを作成する
    nwb = opl.Workbook()
    nws = nwb.active
    with tqdm(total=max_row) as pbar:

        while ws["A" + str(n)].value is not None:
            if ws["A" + str(n)].value == "タイトル":
                nws["A" + str(n + word_amount)].value = ws["A" + str(n)].value
                nws["B" + str(n + word_amount)].value = ws["B" + str(n)].value
                n += 1
                continue
            else:

                ws_value = ws["B" + str(n)].value.replace("「", "『")
                ws_value = ws_value.replace("」", "』")

                length = len(ws_value)
                s_list = []
                s_part1 = []
                s_part2 = []
                s_length = 0
                k = -1
                j = [0]
                count = length // MAX_C
                if count >= 1:
                    word_amount += count
                    # 単語をs_listに代入していく．
                    for token in t.tokenize(ws_value):
                        s_list.append(token.surface)

                    # 単語リストに対応して，品詞名も入れていく．
                    for token in t.tokenize(ws_value):
                        s_part1.append(token.part_of_speech.split(",")[0])

                    # 単語リストに対応して品詞細分類も入れていく．
                    for token in t.tokenize(ws_value):
                        s_part2.append(token.part_of_speech.split(",")[1])

                    for i in range(1, count + 1):

                        while s_length <= int(i * length / (count + 1)) - 6:
                            k += 1
                            s_length += len(s_list[k])

                        while True:
                            try:
                                if s_part2[k] == "括弧開":
                                    j.append(k)
                                    break
                                elif s_part2[k] == "括弧閉":
                                    j.append(k + 1)
                                    break
                                elif s_part1[k] == "記号":
                                    j.append(k + 1)
                                    break
                                elif (
                                    (s_part1[k] == "助詞")
                                    & (s_part1[k + 1] != "助詞")
                                    & (s_part1[k + 1] != "記号")
                                    & (s_part2[k] != "終助詞")
                                ):
                                    j.append(k + 1)
                                    break
                                elif (s_part1[k] == "接続助詞") or (s_part1[k] == "感動詞"):
                                    j.append(k + 1)
                                    break
                                else:
                                    k += 1
                                    s_length += len(s_list[k])
                            except IndexError:
                                break

                    j.append(length)

                    # print(j)

                    for i in range(count + 1):
                        nws["A" + str(n + word_amount - (-i + count))].value = ws[
                            "A" + str(n)
                        ].value
                        nws["B" + str(n + word_amount - (-i + count))].value = "".join(
                            s_list[j[i] : j[i + 1]]
                        )
                        # print("".join(s_list[j[i] : j[i + 1]]))
                        # if len("".join(s_list[j[i] : j[i + 1]])) >= MAX_C + 6:
                        # print("おかしい！")

                    n += 1

                    # pbar.update(1)

                elif count == 0:
                    nws["A" + str(n + word_amount)].value = ws["A" + str(n)].value
                    nws["B" + str(n + word_amount)].value = ws["B" + str(n)].value
                    # print(nws["B"+str(n+l)].value)
                    n += 1

                    # pbar.update(1)

                # print(n+l)

    nwb.save(
        f"{os.path.expanduser('~')}\\works\\{client}\\{name}\\daihon_middle.xlsx"
    )


if __name__ == "__main__":
    name, client = init.main()
    separate(name, client)
