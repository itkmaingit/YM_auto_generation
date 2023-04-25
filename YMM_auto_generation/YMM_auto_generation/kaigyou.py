import getpass
import os
import init
import openpyxl as opl
from janome.tokenizer import Tokenizer
from tqdm import tqdm


def bunkatu(sentence, t):
    """
    文章を1行25文字程度で改行するための関数．tはTokenizeオブジェクトを使用．
    sentenceには改行したい文章を渡す．戻り値は改行された後の文章．
    """

    length = len(sentence)  # 文章自体の長さ．基本75文字以下にしておきたい．
    s_list = []  # 文章の単語リスト
    s_part1 = []  # 各単語の品詞名
    s_part2 = []  # 各単語の品詞細分類(格助詞とか副助詞とか)
    s_length = 0  # 形式定数，改行文字挿入位置判定に使用
    k = -1  # 初めにk+1とするから-1始まりのほうが都合が良い．

    # 単語をs_listに代入していく．
    for token in t.tokenize(sentence):
        s_list.append(token.surface)
    # 単語リストに対応して，品詞名も入れていく．
    for token in t.tokenize(sentence):
        s_part1.append(token.part_of_speech.split(",")[0])
    # 単語リストに対応して品詞細分類も入れていく．
    for token in t.tokenize(sentence):
        s_part2.append(token.part_of_speech.split(",")[1])

    # 文章が15文字以上の時改行
    if 15 <= len(sentence):
        # 文章の半分-3文字から探索．超えた時点で探索終了
        while s_length <= int(length / 2) - 4:
            # s_lengthに各単語ごとに長さを足していく．基準ラインを超えたら足すのをやめて，次のループへ．
            k += 1
            s_length += len(s_list[k])
        # 各条件に従って改行文字を代入していく．
        while True:
            # もし，k番目の単語が"「"だったとき，その一つ手前で問答無用で改行．
            try:
                if s_part2[k] == "括弧開":
                    s_list.insert(k, "\n")
                    break
                # 以下はいい感じに改行文字を代入していく．上と違うのは，"，"などはその1つ後ろで改行したいので，挿入するのは1つ後ろ．
                elif s_part2[k] == "括弧閉":
                    s_list.insert(k + 1, "\n")
                    break
                elif (s_part1[k] == "記号") & (s_part2[k] != "括弧開"):
                    s_list.insert(k + 1, "\n")
                    break
                elif (
                    (s_part1[k] == "助詞")
                    & (s_part1[k + 1] != "助詞")
                    & (s_part1[k + 1] != "記号")
                    & (s_part2[k] != "終助詞")
                ):
                    s_list.insert(k + 1, "\n")
                    break
                elif (s_part1[k] == "接続助詞") or (s_part1[k] == "感動詞"):
                    s_list.insert(k + 1, "\n")
                    break
                else:
                    k += 1
            # どれも条件を満たさなければk+1して，ループの最初へ．
            except IndexError:
                break

        complete_sentence = "".join(s_list)

        return complete_sentence

    else:
        return sentence


def kaigyou(name, client):
    """
    指定したExcelファイルを読み込み，各セルごとに改行を行う関数．
    ただし，セル内に含まれる文字数が50文字以上の場合，文章を2つに分割して改行を行う．

    """
    print("")
    print("---------------kaigyou.py-------------------")

    # 定数設定．
    n = 1  # 開始の列番号
    t = Tokenizer()  # 形態素解析のためのオブジェクト．

    # フォルダ名/VBA_daihon.xlsmを読み込む
    wb = opl.load_workbook(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon_middle.xlsx"
    )
    ws = wb["Sheet"]
    max_row = ws.max_row

    nwb = opl.Workbook()
    nws = nwb.active

    # C(n)のセルが-1になるまで続ける
    with tqdm(total=max_row) as pbar:
        while ws["A" + str(n)].value is not None:
            if ws["A" + str(n)].value == "タイトル" or ws["A" + str(n)].value == "2人":
                nws["A" + str(n)].value = ws["A" + str(n)].value
                nws["B" + str(n)].value = ws["B" + str(n)].value
                n += 1
                continue
            else:
                try:
                    ws_value = ws["B" + str(n)].value.replace("「", "『")
                except AttributeError:
                    pass
                ws_value = ws_value.replace("」", "』")

                nws["A" + str(n)].value = ws["A" + str(n)].value
                try:
                    nws["B" + str(n)].value = bunkatu(ws_value, t)
                except AttributeError:
                    pass

                n += 1
                pbar.update(1)

    # 保存した内容をdaihon_middle.xlsmに保存．
    nwb.save(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon_complete.xlsx"
    )


if __name__ == "__main__":
    name, client = init.main()
    kaigyou(name, client)
