import getpass
from pathlib import Path
import os
import init
import openpyxl as opl
from tqdm import tqdm


def excel(name, client):
    n = 1
    print("")
    print("---------------txt.py-------------------")

    # フォルダ名/txt_fileにEXCELファイルの文章を全て（1行ごとに）書き出す
    dir_path = Path(
        f"{os.path.expanduser('~')}\\works\\" + client + "\\" + name + "\\txt_file"
    )
    dir_path.mkdir(exist_ok=True)

    # フォルダ名/VBA_daihon.xlsmを読み込む
    wb = opl.load_workbook(
        f"{os.path.expanduser('~')}\\works\\"
        + client
        + "\\"
        + name
        + "\\daihon_complete.xlsx"
    )
    ws = wb["Sheet"]

    # 最大列をexcel_sizeとする
    # excel_size = ws.max_row

    # txt_fileにB1の内容をtxtファイルに書き出していく
    while ws["A" + str(n)].value is not None:
        f = open(
            f"{os.path.expanduser('~')}\\works\\"
            + client
            + "\\"
            + name
            + "\\txt_file\\"
            + str(n)
            + ".txt",
            "w",
        )

        try:
            f.write(ws["B" + str(n)].value)

        # 時々，エンコードエラーが発生するのでそれを回避するために例外処理．感情分析には問題なし．
        except UnicodeEncodeError:
            print("\n" + str(n - 1) + "行目がファイル出力をミスった！")

        # 最大列を取得しててもミスる場合はあるのでそのための例外処理。
        # except TypeError:
        #     print("\n" + "終わったようです。")
        #     f.close()
        #     break
        f.close()
        n += 1


if __name__ == "__main__":
    name, client = init.main()
    excel(name, client)
